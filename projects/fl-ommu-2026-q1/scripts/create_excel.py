#!/usr/bin/env python3
"""
Convert CSV files to Excel workbook with multiple sheets.
"""

import csv
import sys

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    print("openpyxl not available, creating simple conversion script")
    sys.exit(1)

def read_csv(filepath):
    """Read CSV file and return as list of lists."""
    with open(filepath, 'r') as f:
        reader = csv.reader(f)
        return list(reader)

def write_sheet(ws, data, sheet_name):
    """Write data to worksheet with formatting."""
    # Write data
    for row_idx, row in enumerate(data, 1):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            
            # Try to convert to number if possible
            if row_idx > 1 and col_idx > 1:  # Skip headers
                try:
                    if '.' in str(value):
                        cell.value = float(value)
                    else:
                        cell.value = int(value)
                except (ValueError, TypeError):
                    cell.value = value
            else:
                cell.value = value
            
            # Format headers
            if row_idx == 1:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
                cell.alignment = Alignment(horizontal='center', wrap_text=True)
    
    # Auto-size columns
    for col in range(1, ws.max_column + 1):
        column_letter = get_column_letter(col)
        max_length = 0
        for cell in ws[column_letter]:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Freeze top row
    ws.freeze_panes = 'A2'

# Create workbook
wb = Workbook()
wb.remove(wb.active)  # Remove default sheet

# Add sheets
sheets = [
    ('Dispensing Locations', 'locations.csv'),
    ('Medical Marijuana (mg THC)', 'mg_thc.csv'),
    ('Low-THC Cannabis (mg CBD)', 'mg_cbd.csv'),
    ('Smoking (oz)', 'smoking_oz.csv'),
    ('Summary Statistics', 'summary.csv'),
    ('All Data Combined', 'all_data.csv'),
]

base_path = '/root/.clawdbot/workspace/ommu_reports_2026_q1/'

for sheet_name, csv_file in sheets:
    print(f"Adding sheet: {sheet_name}")
    data = read_csv(base_path + csv_file)
    ws = wb.create_sheet(title=sheet_name)
    write_sheet(ws, data, sheet_name)

# Save workbook
output_path = '/root/.clawdbot/workspace/FL_MMTC_Q1_2026_Comparison.xlsx'
wb.save(output_path)

print(f"\n✓ Excel workbook created: {output_path}")
print(f"  Total sheets: {len(sheets)}")
for sheet_name, _ in sheets:
    print(f"    - {sheet_name}")
